from __future__ import annotations

import os
import re
from collections.abc import Iterable
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from echosphere.core.test_result import TestResult

INVALID_SHEET_CHARS = r"[\\/*?:\[\]]"
MAX_SHEET_NAME_LEN = 31


class FailedTestExporter:
    """
    Exports failed EchoSphere tests into an Excel workbook. Each failed test gets its own worksheet.
    """

    def __init__(self) -> None:
        """Initialize the exporter state to collect failed TestResult objects."""
        self._results: list[TestResult] = []

    def add_results(self, results: Iterable[TestResult]) -> None:
        """Append only failed TestResult instances from the iterable."""
        for r in results:
            if not r.passed:
                self._results.append(r)

    @staticmethod
    def _sanitize_sheet_name(name: str, used: set[str]) -> str:
        """Return a valid, unique Excel sheet name based on input and used names."""
        sanitized = re.sub(INVALID_SHEET_CHARS, "_", name)
        sanitized = sanitized[:MAX_SHEET_NAME_LEN].strip()
        if not sanitized:
            sanitized = "Sheet"
        base = sanitized
        i = 1
        while sanitized in used:
            suffix = f" ({i})"
            trimmed = base[: MAX_SHEET_NAME_LEN - len(suffix)]
            sanitized = f"{trimmed}{suffix}"
            i += 1
        used.add(sanitized)
        return sanitized

    @staticmethod
    def _auto_fit_columns(ws: Worksheet) -> None:
        """Auto-fit column widths based on cell content."""
        max_width: dict[Any, Any] = {}
        for row in ws.iter_rows(values_only=True):
            for idx, cell in enumerate(row, start=1):
                val = "" if cell is None else str(cell)
                max_width[idx] = max(max_width.get(idx, 0), len(val))
        for idx, width in max_width.items():
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(10, width + 2), 80)

    def write_to_file(self, path: str) -> str:
        """Write failed test details to an Excel (.xlsx) file and return its absolute path."""
        if not path.lower().endswith(".xlsx"):
            path = f"{path}.xlsx"

        abs_path = os.path.abspath(path)
        dir_name = os.path.dirname(abs_path)
        if dir_name and not os.path.exists(dir_name):
            try:
                os.makedirs(dir_name, exist_ok=True)
            except OSError as e:
                raise Exception(f"Failed to create directories for path '{abs_path}': {e}")

        wb = Workbook()
        if self._results:
            default_ws = wb.active
            wb.remove(default_ws)

        used_names: set[str] = set()
        header_font = Font(bold=True)
        meta_label_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for r in self._results:
            sheet_name = self._sanitize_sheet_name(r.name, used_names)
            ws = wb.create_sheet(title=sheet_name)

            meta_rows = [
                ("Test Name:", r.name),
                ("Execution Time:", f"{r.duration:.3f} seconds"),
                ("Failed With:", f"{r.row_count} rows (showing first 1,000)"),
                ("SQL Query:", r.sql),
                ("Execution Timestamp:", r.timestamp.replace(microsecond=0).isoformat()),
            ]
            row_idx = 1
            for label, value in meta_rows:
                ws.cell(row=row_idx, column=1, value=label).font = meta_label_font
                cell = ws.cell(row=row_idx, column=2, value=value)
                cell.alignment = wrap_alignment
                row_idx += 1

            row_idx += 1  # blank line

            columns = r.failure_columns or []
            rows = r.failure_rows or []

            for col_idx, col_name in enumerate(columns, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=col_name)
                c.font = header_font
            header_row = row_idx

            for dr in rows:
                row_idx += 1
                for col_idx, cell_val in enumerate(dr, start=1):
                    if isinstance(cell_val, datetime):
                        cell_val.replace(tzinfo=None)
                    ws.cell(row=row_idx, column=col_idx, value=cell_val)

            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            self._auto_fit_columns(ws)

        if not self._results:
            ws = wb.active
            ws.title = "No Failures"
            ws.cell(row=1, column=1, value="No failed tests to export.")

        try:
            wb.save(abs_path)
        except OSError as e:
            raise Exception(f"Failed to write Excel file to '{abs_path}': {e}")

        return abs_path
